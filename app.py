import streamlit as st
import os
import io
import shutil
import pandas as pd
from PIL import Image, ImageDraw, ImageFilter
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import textwrap
from dataclasses import dataclass
from typing import Optional, Tuple, List, Dict
import tempfile
import PyPDF2

@dataclass
class PDFConfig:
    """Configuration class for PDF layout and styling"""
    # Page and grid settings
    page_width: float = A4[0]
    page_height: float = A4[1]
    margin: int = 30
    grid_cols: int = 3
    grid_rows: int = 4
    
    # --- EDITED CODE BLOCK ---
    # These are now fallbacks. The actual font will be set to Poppins if available.
    font_heading: str = "Helvetica-Bold"
    font_text: str = "Helvetica"
    # --- END EDITED CODE BLOCK ---

    font_size_heading: int = 24
    font_size_category: int = 18
    font_size_name: int = 10
    font_size_price: int = 12
    font_size_index_title: int = 32
    font_size_index_category: int = 16
    font_size_page_number: int = 10



    PRODUCT_NAME_FONT_SIZE = 10
    PRODUCT_NAME_LINE_HEIGHT = 13
    PRODUCT_NAME_MAX_LINES = 2
    IMAGE_TEXT_PADDING = 25


    NAME_FONT_SIZE = 10
    NAME_LINE_HEIGHT = 10
    MAX_NAME_LINES = 2
    IMAGE_NAME_GAP = 30


    
    image_dpi: int = 300
    image_quality: int = 95
    
    category_colors: List[dict] = None
    color_text: colors.Color = colors.HexColor("#2C3E50")
    color_price: colors.Color = colors.HexColor("#27AE60")
    color_index_bg: colors.Color = colors.HexColor("#FFFFFF")
    
    def __post_init__(self):
        if self.category_colors is None:
            self.category_colors = [
                {'name': 'Lilac Dreams', 'start': colors.HexColor("#E8D5FF"), 'end': colors.HexColor("#F5F0FF"), 'accent': colors.HexColor("#9B59B6")},
                {'name': 'Ocean Breeze', 'start': colors.HexColor("#D4E6FF"), 'end': colors.HexColor("#F0F6FF"), 'accent': colors.HexColor("#3498DB")},
                {'name': 'Mint Fresh', 'start': colors.HexColor("#D5FFE0"), 'end': colors.HexColor("#F0FFF4"), 'accent': colors.HexColor("#2ECC71")},
                {'name': 'Rose Glow', 'start': colors.HexColor("#FFE0E6"), 'end': colors.HexColor("#FFF5F7"), 'accent': colors.HexColor("#E74C3C")},
                {'name': 'Golden Hour', 'start': colors.HexColor("#FFF2D5"), 'end': colors.HexColor("#FFFBF0"), 'accent': colors.HexColor("#F39C12")},
                {'name': 'Lavender Mist', 'start': colors.HexColor("#E6E0FF"), 'end': colors.HexColor("#F7F5FF"), 'accent': colors.HexColor("#8E44AD")}
            ]
    
    @property
    def cell_width(self) -> float: return (self.page_width - 2 * self.margin) / self.grid_cols
    @property
    def cell_height(self) -> float: return (self.page_height - 120) / self.grid_rows
    @property
    def image_height(self) -> float: return self.cell_height * 0.6
    @property
    def image_width(self) -> float: return self.cell_width - 20

class ImageProcessor:
    """Handles image extraction and processing from Excel files with enhanced quality"""
    
    def __init__(self, image_dir: str = "images"):
        self.image_dir = image_dir
        self._ensure_image_dir()
    
    def _ensure_image_dir(self):
        if not os.path.exists(self.image_dir):
            os.makedirs(self.image_dir)
    
    def create_placeholder_image(self, width: int = 600, height: int = 600) -> str:
        placeholder = Image.new("RGB", (width, height), color="#F8F9FA")
        draw = ImageDraw.Draw(placeholder)
        draw.rectangle([0, 0, width-1, height-1], outline="#E0E0E0", width=3)
        center_x, center_y = width // 2, height // 2
        icon_size = min(width, height) // 4
        draw.rectangle([center_x - icon_size, center_y - icon_size//2, center_x + icon_size, center_y + icon_size//2], fill="#BDC3C7", outline="#95A5A6", width=2)
        try:
            from PIL import ImageFont
            font = ImageFont.load_default()
        except:
            font = None
        text = "No Image"
        if font:
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
        else:
            text_width = len(text) * 12
        text_x = center_x - text_width // 2
        text_y = center_y + icon_size
        draw.text((text_x, text_y), text, fill="#7F8C8D", font=font)
        placeholder_path = os.path.join(self.image_dir, "placeholder.jpg")
        placeholder.save(placeholder_path, "JPEG", quality=95, optimize=True)
        return placeholder_path
    
    def _enhance_image_quality(self, image: Image.Image, target_width: int, target_height: int) -> Image.Image:
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode in ('RGBA', 'LA') else None)
            image = background
        elif image.mode != 'RGB':
            image = image.convert('RGB')
        
        original_width, original_height = image.size
        aspect_ratio = original_width / original_height
        target_aspect_ratio = target_width / target_height
        
        if aspect_ratio > target_aspect_ratio:
            new_width = target_width
            new_height = int(target_width / aspect_ratio)
        else:
            new_height = target_height
            new_width = int(target_height * aspect_ratio)
        
        resized_image = image.resize((new_width, new_height), Image.Resampling.LANCZOS)
        resized_image = resized_image.filter(ImageFilter.UnsharpMask(radius=1, percent=150, threshold=3))
        
        final_image = Image.new('RGB', (target_width, target_height), (255, 255, 255))
        x_offset = (target_width - new_width) // 2
        y_offset = (target_height - new_height) // 2
        final_image.paste(resized_image, (x_offset, y_offset))
        
        return final_image
    
    def extract_images_from_excel(self, excel_file: str, data_length: int, start_row: int = 2) -> list:
        wb = load_workbook(excel_file)
        sheet = wb['Sheet1']
        image_loader = SheetImageLoader(sheet)
        image_filenames = []
        for i in range(data_length):
            row_num = start_row + i
            cell = f'C{row_num}'
            try:
                image = image_loader.get(cell)
                if image:
                    enhanced_image = self._enhance_image_quality(image, 600, 600)
                    image_filename = f"image_{row_num}.jpg"
                    image_path = os.path.join(self.image_dir, image_filename)
                    enhanced_image.save(image_path, "JPEG", quality=95, optimize=True, dpi=(300, 300))
                    image_filenames.append(image_filename)
                else:
                    image_filenames.append("")
            except Exception as e:
                st.warning(f"Could not extract image from cell {cell}: {e}")
                image_filenames.append("")
        return image_filenames
    
    def cleanup(self):
        if os.path.exists(self.image_dir):
            shutil.rmtree(self.image_dir, ignore_errors=True)

class DataProcessor:
    @staticmethod
    def load_excel_data(excel_file: str, start_row: int = 2, end_row: Optional[int] = None) -> pd.DataFrame:
        wb = load_workbook(excel_file)
        sheet = wb['Sheet1']
        data = pd.DataFrame(sheet.values)
        headers = data.iloc[0]
        data = data[1:]
        data.columns = headers
        data = data.dropna(how='all').reset_index(drop=True)
        start_idx = max(0, start_row - 2)
        if end_row is not None:
            end_idx = min(len(data), end_row - 1)
            data = data.iloc[start_idx:end_idx]
        else:
            data = data.iloc[start_idx:]
        data = data.reset_index(drop=True)
        return data
    
    @staticmethod
    def process_categories(data: pd.DataFrame) -> List[dict]:
        categories = []
        current_category = None
        current_products = []
        for idx, row in data.iterrows():
            name = row.get("Name", "")
            cat = row.get("CAT", "")
            if pd.isna(name) or str(name).strip() == "" or str(name).lower() == "nan":
                if current_category and current_products:
                    categories.append({'name': current_category, 'products': current_products.copy()})
                current_category = None
                current_products = []
                continue
            if pd.notna(cat) and str(cat).strip() != "" and str(cat).lower() != "nan":
                if current_category and current_products:
                    categories.append({'name': current_category, 'products': current_products.copy()})
                current_category = str(cat).strip()
                current_products = []
            if current_category:
                product_data = row.copy()
                product_data['original_index'] = idx
                current_products.append(product_data)
        if current_category and current_products:
            categories.append({'name': current_category, 'products': current_products.copy()})
        return categories
    
    @staticmethod
    def validate_data(data: pd.DataFrame) -> Tuple[bool, str]:
        required_columns = ["Name", "Price E", "CAT"]
        missing_columns = [col for col in required_columns if col not in data.columns]
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"
        if data.empty:
            return False, "No data found in the specified range"
        return True, "Data validation successful"

class PDFGenerator:
    """Handles PDF generation with enhanced font rendering and image quality"""
    
    # --- EDITED CODE BLOCK START ---
    def __init__(self, config: PDFConfig):
        self.config = config
        self.bookmarks = {}
        self.total_pages = 0

        # Register custom fonts that support the Rupee symbol (₹).
        # This is the crucial step to fix the "green square" issue.
        # Make sure 'Poppins-Regular.ttf' and 'Poppins-Bold.ttf' are in the same folder.
        try:
            pdfmetrics.registerFont(TTFont('Poppins-Regular', 'Poppins-Regular.ttf'))
            pdfmetrics.registerFont(TTFont('Poppins-Bold', 'Poppins-Bold.ttf'))
            
            # Update the config to use these new, better fonts
            self.config.font_text = 'Poppins-Regular'
            self.config.font_heading = 'Poppins-Bold'
            st.info("✅ Poppins font loaded successfully. Rupee symbol will be displayed.")
        except Exception:
            st.warning(
                "⚠️ Poppins font files not found. The Rupee symbol (₹) may not display correctly. "
                "Please download 'Poppins-Regular.ttf' and 'Poppins-Bold.ttf' and place them "
                "in the same directory as the script."
            )
            # Fallback to default fonts if Poppins is not found
            self.config.font_text = 'Helvetica'
            self.config.font_heading = 'Helvetica-Bold'
    
    
    def _wrap_text(self, text: str, width: int) -> str:
        return textwrap.fill(str(text), width)
    

    def _draw_centered_name_block(
        self,
        canvas_obj,
        x,
        image_bottom_y,
        price_top_y,
        lines,
    ):
        font_size = 10
        line_height = 13

        canvas_obj.setFont(self.config.font_text, font_size)
        canvas_obj.setFillColor(self.config.color_text)

        block_height = len(lines) * line_height
        available_height = image_bottom_y - price_top_y

        # Center name block between image and price
        start_y = price_top_y + (available_height - block_height) / 2 + block_height

        for i, line in enumerate(lines):
            canvas_obj.drawString(
                x + 12,
                start_y - i * line_height,
                line
            )

    
    def _draw_gradient_background(self, canvas_obj: canvas.Canvas, color_scheme: dict):
        num_steps = 30
        height_step = self.config.page_height / num_steps
        start_color = color_scheme['start']
        end_color = color_scheme['end']
        for i in range(num_steps):
            ratio = i / (num_steps - 1)
            r = start_color.red + (end_color.red - start_color.red) * ratio
            g = start_color.green + (end_color.green - start_color.green) * ratio
            b = start_color.blue + (end_color.blue - start_color.blue) * ratio
            interpolated_color = colors.Color(r, g, b)
            canvas_obj.setFillColor(interpolated_color)
            y = self.config.page_height - (i + 1) * height_step
            canvas_obj.rect(0, y, self.config.page_width, height_step, fill=True, stroke=False)
    
    def _draw_index_background(self, canvas_obj: canvas.Canvas):
        canvas_obj.setFillColor(self.config.color_index_bg)
        canvas_obj.rect(0, 0, self.config.page_width, self.config.page_height, fill=True, stroke=False)
        canvas_obj.setStrokeColor(colors.HexColor("#BDC3C7"))
        canvas_obj.setLineWidth(2)
        canvas_obj.line(self.config.margin, self.config.page_height - 80, self.config.page_width - self.config.margin, self.config.page_height - 80)
    
    def _draw_header(self, canvas_obj: canvas.Canvas, category_name: str, color_scheme: dict):
        canvas_obj.setFont(self.config.font_heading, self.config.font_size_category)
        canvas_obj.setFillColor(color_scheme['accent'])
        canvas_obj.drawCentredString(self.config.page_width / 2, self.config.page_height - 50, category_name)
        canvas_obj.setStrokeColor(color_scheme['accent'])
        canvas_obj.setLineWidth(3)
        line_width = pdfmetrics.stringWidth(category_name, self.config.font_heading, self.config.font_size_category) + 20
        start_x = (self.config.page_width - line_width) / 2
        canvas_obj.line(start_x, self.config.page_height - 65, start_x + line_width, self.config.page_height - 65)
    
    def _draw_page_number(self, canvas_obj: canvas.Canvas, page_num: int):
        canvas_obj.setFont(self.config.font_text, self.config.font_size_page_number)
        canvas_obj.setFillColor(colors.HexColor("#666666"))
        canvas_obj.drawCentredString(self.config.page_width / 2, 20, f"Page {page_num}")
    
    def _create_index_page(self, canvas_obj: canvas.Canvas, categories: List[dict]):
        max_items_per_page = int((self.config.page_height - 260) / 40)
        total_pages = (len(categories) + max_items_per_page - 1) // max_items_per_page

        category_chunks = [
            categories[i:i + max_items_per_page]
            for i in range(0, len(categories), max_items_per_page)
        ]

        for page_idx, chunk in enumerate(category_chunks):
            if page_idx > 0:
                canvas_obj.showPage()

            self._draw_index_background(canvas_obj)

            # Title
            canvas_obj.setFont(self.config.font_heading, self.config.font_size_index_title)
            canvas_obj.setFillColor(colors.HexColor("#2C3E50"))
            canvas_obj.drawCentredString(
                self.config.page_width / 2,
                self.config.page_height - 60,
                "Index"
            )

            start_y = self.config.page_height - 140
            line_height = 40

            for idx, category in enumerate(chunk):
                global_idx = page_idx * max_items_per_page + idx
                category_name = category['name']
                page_number = self.bookmarks.get(category_name, 1)
                color_scheme = self.config.category_colors[global_idx % len(self.config.category_colors)]
                current_y = start_y - (idx * line_height)

                # Background row
                canvas_obj.setFillColor(colors.Color(0.97, 0.97, 0.97))
                canvas_obj.roundRect(
                    self.config.margin,
                    current_y - 12,
                    self.config.page_width - 2 * self.config.margin,
                    30,
                    8,
                    fill=True,
                    stroke=False
                )

                # Index number
                canvas_obj.setFillColor(color_scheme['accent'])
                canvas_obj.setFont(self.config.font_heading, self.config.font_size_index_category)
                canvas_obj.drawString(self.config.margin + 25, current_y, f"{global_idx + 1}.")

                # Category name + link
                canvas_obj.setFont(self.config.font_text, self.config.font_size_index_category)
                link_rect = (
                    self.config.margin + 60,
                    current_y - 8,
                    self.config.page_width - self.config.margin - 100,
                    current_y + 20
                )
                canvas_obj.linkAbsolute("", f"category_{global_idx}", link_rect)
                canvas_obj.drawString(self.config.margin + 60, current_y, category_name)

                # Dotted line
                text_width = pdfmetrics.stringWidth(
                    category_name,
                    self.config.font_text,
                    self.config.font_size_index_category
                )
                dots_start_x = self.config.margin + 60 + text_width + 10
                dots_end_x = self.config.page_width - self.config.margin - 60

                canvas_obj.setStrokeColor(colors.HexColor("#BDC3C7"))
                canvas_obj.setDash(3, 4)
                canvas_obj.line(dots_start_x, current_y + 7, dots_end_x, current_y + 7)
                canvas_obj.setDash()

                # Page number
                canvas_obj.setFillColor(colors.HexColor("#2C3E50"))
                canvas_obj.setFont(self.config.font_heading, self.config.font_size_index_category)
                canvas_obj.drawRightString(
                    self.config.page_width - self.config.margin - 25,
                    current_y,
                    str(page_number)
                )

            # Footer
            canvas_obj.setFont(self.config.font_text, 11)
            canvas_obj.setFillColor(colors.HexColor("#95A5A6"))
            canvas_obj.drawCentredString(
                self.config.page_width / 2,
                120,
                "Click on any category name to navigate directly to that section"
            )

    def _wrap_product_name(self, text: str) -> list[str]:
        lines = textwrap.wrap(str(text), width=28)
        if len(lines) > self.config.MAX_NAME_LINES:
            lines = lines[:self.config.MAX_NAME_LINES]
            lines[-1] = lines[-1][:-3] + "..."
        return lines



    
    def _draw_product(self, canvas_obj: canvas.Canvas, product: pd.Series, x: float, y: float, placeholder_image: str, temp_dir: str, color_scheme: dict):
        product_name = str(product.get("Name", "Unknown Product"))
        
        # This part was already correct from your previous request
        price_value = product.get("Price E", "Unknown Price")
        try:
            formatted_price = str(int(float(price_value)))
        except (ValueError, TypeError):
            formatted_price = str(price_value)
        
        image_name = product.get("Image Name", None)
        canvas_obj.setFillColor(colors.Color(1, 1, 1, alpha=0.8))
        canvas_obj.roundRect(x + 5, y - self.config.cell_height + 5, self.config.cell_width - 10, self.config.cell_height - 10, 12, fill=True, stroke=False)
        canvas_obj.setStrokeColor(color_scheme['accent'])
        canvas_obj.setLineWidth(0.8)
        canvas_obj.roundRect(x + 5, y - self.config.cell_height + 5, self.config.cell_width - 10, self.config.cell_height - 10, 12, fill=False, stroke=True)
        
        try:
            image_path = os.path.join("images", image_name) if image_name else placeholder_image
            if not os.path.exists(image_path): image_path = placeholder_image
            with Image.open(image_path) as img:
                img_width, img_height = img.size
            
            target_width, target_height = self.config.image_width, self.config.image_height
            aspect_ratio, target_aspect_ratio = img_width / img_height, target_width / target_height
            
            if aspect_ratio > target_aspect_ratio:
                draw_width, draw_height = target_width, target_width / aspect_ratio
            else:
                draw_height, draw_width = target_height, target_height * aspect_ratio
            
            image_x = x + (self.config.cell_width - draw_width) / 2
            image_y = y - self.config.image_height - 15 + (target_height - draw_height) / 2
            
            canvas_obj.drawImage(image_path, image_x, image_y, width=draw_width, height=draw_height, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            st.warning(f"Error loading image for {product_name}: {e}")
    
        name_lines = self._wrap_product_name(product_name)

        canvas_obj.setFont(self.config.font_text, self.config.NAME_FONT_SIZE)
        canvas_obj.setFillColor(self.config.color_text)

        # Base Y = first line position if 2-line name
        first_line_y = (
            y
            - self.config.image_height
            - self.config.IMAGE_NAME_GAP
        )

        if len(name_lines) == 1:
            # Center the single line within a 2-line area
            name_y = first_line_y - (self.config.NAME_LINE_HEIGHT / 2)
            canvas_obj.drawString(x + 12, name_y, name_lines[0])

        else:
            # Exactly 25px below image, no centering
            canvas_obj.drawString(x + 12, first_line_y, name_lines[0])
            canvas_obj.drawString(x + 12, first_line_y - self.config.NAME_LINE_HEIGHT, name_lines[1])

        
        price_y = y - self.config.cell_height + 17
        # canvas_obj.setFont(self.config.font_heading, self.config.font_size_price)
        # canvas_obj.setFillColor(self.config.color_price)
        
        # canvas_obj.drawString(x + 10, price_y, f"₹ {formatted_price}")


        canvas_obj.setFont(self.config.font_heading, self.config.font_size_price)
        canvas_obj.setFillColor(self.config.color_price)

        price_value = product.get("Price E", None)

        if price_value is None or str(price_value).strip() == "" or str(price_value).lower() == "nan":
            # Price missing: show placeholder without ₹
            display_price = "Ask"
        else:
            try:
                display_price = f"₹ {int(float(price_value))}"
            except (ValueError, TypeError):
                display_price = str(price_value)

        canvas_obj.drawString(x + 10, price_y, display_price)

    def create_pdf(self, categories: List[dict], placeholder_image: str, output_pdf: str) -> str:
        c = canvas.Canvas(output_pdf, pagesize=A4)
        items_per_page = self.config.grid_rows * self.config.grid_cols
        
        current_page = 2
        for category in categories:
            self.bookmarks[category['name']] = current_page
            total_products = len(category['products'])
            pages_needed = (total_products + items_per_page - 1) // items_per_page if total_products > 0 else 1
            current_page += pages_needed
        
        self.total_pages = current_page - 1
        self._create_index_page(c, categories)
        
        with tempfile.TemporaryDirectory() as temp_dir:
            current_page = 2
            for cat_idx, category in enumerate(categories):
                category_name = category['name']
                products = category['products']
                color_scheme = self.config.category_colors[cat_idx % len(self.config.category_colors)]
                total_products = len(products)
                pages_needed = (total_products + items_per_page - 1) // items_per_page if total_products > 0 else 1
                
                for page_num in range(pages_needed):
                    c.showPage()
                    if page_num == 0: c.bookmarkPage(f"category_{cat_idx}")
                    
                    self._draw_gradient_background(c, color_scheme)
                    self._draw_header(c, category_name, color_scheme)
                    # self._draw_page_number(c, current_page) # Optional page numbers
                    
                    start_idx = page_num * items_per_page
                    end_idx = min((page_num + 1) * items_per_page, total_products)
                    page_products = products[start_idx:end_idx]
                    
                    grid_start_y = self.config.page_height - 100
                    for idx, product in enumerate(page_products):
                        current_row = idx // self.config.grid_cols
                        current_col = idx % self.config.grid_cols
                        x = self.config.margin + current_col * self.config.cell_width
                        y = grid_start_y - current_row * self.config.cell_height
                        self._draw_product(c, product, x, y, placeholder_image, temp_dir, color_scheme)
                    
                    current_page += 1
        c.save()
        return output_pdf

class CatalogGenerator:
    def __init__(self):
        self.config = PDFConfig()
        self.image_processor = ImageProcessor()
        self.data_processor = DataProcessor()
        self.pdf_generator = PDFGenerator(self.config)
    
    def generate_catalog(self, excel_file: str, start_row: int = 2, end_row: Optional[int] = None) -> bytes:
        temp_catalog_path = None
        merged_pdf_path = None
        try:
            data = self.data_processor.load_excel_data(excel_file, start_row, end_row)
            is_valid, message = self.data_processor.validate_data(data)
            if not is_valid: raise ValueError(message)
            
            categories = self.data_processor.process_categories(data)
            if not categories: raise ValueError("No valid categories found in the data")
            
            st.info(f"Processing {len(categories)} categories...")
            
            if any(cat['products'] for cat in categories):
                image_filenames = self.image_processor.extract_images_from_excel(excel_file, len(data), start_row)
                for category in categories:
                    for product in category['products']:
                        original_idx = product.get('original_index', 0)
                        if original_idx < len(image_filenames):
                            product['Image Name'] = image_filenames[original_idx]
                        else:
                            product['Image Name'] = ""
            
            placeholder_path = self.image_processor.create_placeholder_image()
            
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
                temp_catalog_path = temp_pdf.name
            
            self.pdf_generator.create_pdf(categories, placeholder_path, temp_catalog_path)
            
            first_pdf_to_merge = "first.pdf"
            final_pdf_path = temp_catalog_path
            if os.path.exists(first_pdf_to_merge):
                st.info(f"Found '{first_pdf_to_merge}', merging it at the beginning of the catalog.")
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_merged_pdf:
                    merged_pdf_path = temp_merged_pdf.name
                
                merger = PyPDF2.PdfWriter()
                merger.append(first_pdf_to_merge)
                merger.append(temp_catalog_path)
                with open(merged_pdf_path, "wb") as f_out:
                    merger.write(f_out)
                merger.close()
                final_pdf_path = merged_pdf_path
            else:
                st.warning(f"'{first_pdf_to_merge}' not found. Skipping merge.")

            with open(final_pdf_path, "rb") as f:
                pdf_bytes = f.read()
            return pdf_bytes
        except Exception as e:
            raise Exception(f"Error generating catalog: {str(e)}")
        finally:
            self.image_processor.cleanup()
            if temp_catalog_path and os.path.exists(temp_catalog_path): os.unlink(temp_catalog_path)
            if merged_pdf_path and os.path.exists(merged_pdf_path): os.unlink(merged_pdf_path)

def main():
    st.title("📚 Product Catalog Generator with Interactive Index")
    st.markdown("Generate professional product catalogs from Excel files with clickable table of contents")
    
    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx'], help="Excel file should have columns: Name, Price E, CAT, and images in column C")
    
    if uploaded_file:
        with st.expander("⚙️ Advanced Settings", expanded=False):
            st.markdown("**Row Range Selection**")
            col1, col2 = st.columns(2)
            with col1:
                start_row = st.number_input("Start Row", min_value=2, value=2, help="First row to include (row 1 is headers)")
            with col2:
                use_end_row = st.checkbox("Set End Row")
                end_row = None
                if use_end_row:
                    end_row = st.number_input("End Row", min_value=start_row, value=start_row + 10, help="Last row to include")
        
        if st.button("🚀 Generate PDF Catalog with Index", type="primary"):
            try:
                with st.spinner("Processing Excel file and generating catalog..."):
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_file.write(uploaded_file.getvalue())
                        temp_excel_path = temp_file.name
                    
                    generator = CatalogGenerator()
                    pdf_bytes = generator.generate_catalog(temp_excel_path, start_row, end_row)
                    os.unlink(temp_excel_path)
                
                st.success("✅ PDF catalog with interactive index generated successfully!")
                row_info = f"Rows {start_row}" + (f" to {end_row}" if end_row else " to end")
                st.info(f"📊 Generated catalog for {row_info} with clickable table of contents")
                st.download_button(label="📥 Download Interactive PDF Catalog", data=pdf_bytes, file_name=f"interactive_catalog_rows_{start_row}{'_to_' + str(end_row) if end_row else '_to_end'}.pdf", mime="application/pdf", type="primary")
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                st.info("Please check your Excel file format and try again.")
    else:
        st.info("👆 Please upload an Excel file to get started")
        with st.expander("📋 Instructions", expanded=True):
            st.markdown("""
            **File Format Requirements:**
            - Excel file (.xlsx format)
            - Column structure: **Name**, **Price E**, **CAT**, and images in **Column C**
            
            **New Features:**
            - 📑 **Interactive Index Page**: First page contains a clickable table of contents.
            - 🔗 **Clickable Navigation**: Click any category name to jump directly to that section.
            - 🎨 **Enhanced Design**: Professional layout with color-coded categories.
            - 📖 **PDF Bookmarks**: For easy navigation in PDF viewers.
            
            **Features:**
            - If a file named `first.pdf` is in the same directory, it will be added to the start of the catalog.
            - Category-based organization with beautiful gradients.
            - Automatic image extraction from Excel.
            - Professional PDF layout (3x4 grid per page).
            - NULL name detection for category separation.
            
            **Tips:**
            - Leave Name cell empty to start a new category.
            - Only the first product in each category needs the CAT value.
            """)

if __name__ == "__main__":
    main()